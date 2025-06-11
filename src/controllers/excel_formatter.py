"""
Excel formatter controller for XRF Data Manager.
Handles formatting and saving tables to Excel.
"""

import os
import sys
from typing import Dict, Any, List

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))


def save_tables_to_excel(
    tables: Dict[str, pd.DataFrame],
    excel_path: str,
    metadata: Dict[str, Any],
    missing_data: str = '---'
) -> bool:
    """
    Save tables to Excel with formatting.
    
    Args:
        tables: Dictionary of table DataFrames
        excel_path: Path to save Excel file
        metadata: Project metadata
        missing_data: String to use for missing data
        
    Returns:
        True if saved successfully, False otherwise
    """
    if not tables:
        return False
    
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb.get_sheet_by_name('Sheet'))
        
        # Add metadata and lookup sheets first
        if 'metadata' in tables:
            create_metadata_sheet(wb, tables['metadata'], 'Metadata')
            # Remove from tables dict to avoid processing again
            del tables['metadata']
            
        if 'lookup' in tables:
            create_lookup_sheet(wb, tables['lookup'], 'Lookup Table')
            # Remove from tables dict to avoid processing again
            del tables['lookup']
        
        # Create a sheet for each table
        for table_name, df in tables.items():
            # Create sheet with descriptive name
            sheet_name = format_sheet_name(table_name)
            ws = wb.create_sheet(sheet_name)
            
            # Create caption
            caption = create_caption(table_name, metadata)
            ws.append([caption])
            ws.append([])  # Empty row after caption
            
            # Fill NaN with missing data string
            df_clean = df.copy()
            df_clean = df_clean.fillna(missing_data)
            
            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=True)):
                ws.append(row)
            
            # Apply formatting
            format_excel_sheet(ws, len(df_clean.columns), len(df_clean)+1)  # +1 for header
        
        # Save workbook
        wb.save(excel_path)
        return True
    
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")
        return False


def create_metadata_sheet(wb, metadata_df, sheet_name):
    """
    Create a sheet with metadata information.
    
    Args:
        wb: Workbook
        metadata_df: Metadata DataFrame
        sheet_name: Sheet name
    """
    ws = wb.create_sheet(sheet_name)
    
    # Create caption
    ws.append(["Project Metadata"])
    ws.append([])  # Empty row after caption
    
    # Convert metadata to a more readable format
    readable_metadata = []
    for column in metadata_df.columns:
        readable_metadata.append([column, str(metadata_df[column].iloc[0])])
    
    # Write metadata
    for row in readable_metadata:
        ws.append(row)
    
    # Apply formatting
    format_metadata_sheet(ws, 2, len(readable_metadata))


def create_lookup_sheet(wb, lookup_df, sheet_name):
    """
    Create a sheet with lookup table information.
    
    Args:
        wb: Workbook
        lookup_df: Lookup DataFrame
        sheet_name: Sheet name
    """
    ws = wb.create_sheet(sheet_name)
    
    # Create caption
    ws.append(["Sample Lookup Table"])
    ws.append([])  # Empty row after caption
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(lookup_df, index=False, header=True)):
        ws.append(row)
    
    # Apply formatting
    format_excel_sheet(ws, len(lookup_df.columns), len(lookup_df)+1)  # +1 for header


def format_metadata_sheet(worksheet, num_cols, num_rows):
    """
    Apply Excel formatting to metadata worksheet.
    
    Args:
        worksheet: openpyxl worksheet
        num_cols: Number of columns
        num_rows: Number of rows
    """
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply caption formatting
    caption_cell = worksheet.cell(row=1, column=1)
    caption_cell.font = Font(name='Arial', size=12, bold=True)
    caption_cell.alignment = Alignment(horizontal='left')
    
    # Get the range of data cells (excluding caption and empty row)
    data_start_row = 3
    data_end_row = data_start_row + num_rows
    
    # Format data rows
    for row in range(data_start_row, data_end_row):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            if col == 1:  # Key column
                cell.font = Font(name='Arial', size=10, bold=True)
            else:  # Value column
                cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
    
    # Auto-size columns
    for col in range(1, num_cols + 1):
        column_letter = worksheet.cell(row=1, column=col).column_letter
        worksheet.column_dimensions[column_letter].auto_size = True


def format_sheet_name(table_name: str) -> str:
    """
    Format table name to sheet name.
    
    Args:
        table_name: Original table name
        
    Returns:
        Formatted sheet name
    """
    # Split by underscore and capitalize
    parts = table_name.split('_')
    name = ' '.join(part.capitalize() for part in parts)
    
    # Map specific names
    name_map = {
        'Absolute Major Elements': 'Absolute Major Elements',
        'Absolute Major Oxides': 'Absolute Major Oxides',
        'Absolute Trace Elements': 'Absolute Trace Elements',
        'Absolute Trace Oxides': 'Absolute Trace Oxides',
        'Relative Major Elements': 'Relative Major Elements',
        'Relative Major Oxides': 'Relative Major Oxides',
        'Relative Trace Elements': 'Relative Trace Elements',
        'Relative Trace Oxides': 'Relative Trace Oxides'
    }
    
    return name_map.get(name, name)


def create_caption(table_name: str, metadata: Dict[str, Any]) -> str:
    """
    Create a caption for the table.
    
    Args:
        table_name: Table name
        metadata: Project metadata
        
    Returns:
        Caption string
    """
    # Get metadata
    project_number = metadata.get('project_number', '')
    project_name = metadata.get('project_name', '')
    client_name = metadata.get('client_name', '')
    
    # Create base caption
    caption = f"Table X. "
    
    # Add table type specific text with units
    if 'absolute_major' in table_name:
        caption += "Absolute major element concentrations"
        if 'element' in table_name:
            caption += " (wt.%)"
    elif 'absolute_trace' in table_name:
        caption += "Absolute trace element concentrations"
        if 'element' in table_name:
            caption += " (ppm)"
    elif 'relative_major' in table_name:
        caption += "Relative major element concentrations"
        if 'element' in table_name:
            caption += " (wt.%)"
    elif 'relative_trace' in table_name:
        caption += "Relative trace element concentrations"
        if 'element' in table_name:
            caption += " (ppm)"
    
    # Add oxide info if in table name
    if 'oxide' in table_name:
        if 'major' in table_name:
            caption += " reported as oxides (wt.%)"
        else:  # trace
            caption += " reported as oxides (ppm)"
    
    # Add project info
    if project_number or project_name:
        caption += f" for {project_number}"
        if project_name:
            caption += f" {project_name}"
    
    # Add client info
    if client_name:
        caption += f" ({client_name})"
    
    return caption


def format_excel_sheet(worksheet, num_cols, num_rows):
    """
    Apply Excel formatting to worksheet.
    
    Args:
        worksheet: openpyxl worksheet
        num_cols: Number of columns
        num_rows: Number of rows
    """
    # Define styles
    header_font = Font(name='Arial', size=10, bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    normal_font = Font(name='Arial', size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_bottom_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='double')
    )
    
    # Apply caption formatting
    caption_cell = worksheet.cell(row=1, column=1)
    caption_cell.font = Font(name='Arial', size=12, italic=False)
    caption_cell.alignment = Alignment(horizontal='left')
    
    # Get the range of data cells (excluding caption and empty row)
    data_start_row = 3
    data_end_row = data_start_row + num_rows
    
    # Format header row
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=data_start_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Find total row index (if exists)
    total_row_index = None
    for row in range(data_start_row + 1, data_end_row):
        cell_value = worksheet.cell(row=row, column=2).value
        if cell_value == 'Total':
            total_row_index = row
            break
    
    # Format data rows
    for row in range(data_start_row + 1, data_end_row):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = normal_font
            
            # Apply appropriate border - double line ABOVE the total row
            if total_row_index and row == total_row_index - 1:
                # Cell above total row gets double bottom border
                cell.border = thick_bottom_border
            else:
                cell.border = thin_border
            
            # Center align all cells
            cell.alignment = Alignment(horizontal='center')
    
    # Auto-size columns
    for col in range(1, num_cols + 1):
        column_letter = worksheet.cell(row=1, column=col).column_letter
        worksheet.column_dimensions[column_letter].auto_size = True
